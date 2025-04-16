import os
import json
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF, HTMLMixin
import xlsxwriter
from jinja2 import Template, Environment, FileSystemLoader
import plotly.graph_objects as go
import plotly.io as pio
import plotly.express as px
import base64
from io import BytesIO
import pdfkit
from openpyxl import Workbook
import openpyxl.styles
import openpyxl.utils
import math

class PremiumPDF(FPDF):
    """Premium PDF class with enhanced styling and formatting capabilities"""
    
    def __init__(self, orientation='P', unit='mm', format='A4'):
        super().__init__(orientation, unit, format)
        # Live Oak Bank color palette
        self.primary_color = (30, 24, 71)     # Dark blue/purple #1E1847
        self.primary_light = (232, 231, 240)  # Light blue/purple #e8e7f0
        self.accent_color = (45, 199, 122)    # Green #2DC77A
        self.accent_light = (233, 249, 242)   # Light green #e9f9f2
        self.danger_color = (239, 68, 68)     # Red #ef4444
        self.warning_color = (245, 158, 11)   # Orange #f59e0b
        self.success_color = (45, 199, 122)   # Green #2DC77A
        self.links = {}  # For storing link destinations
        
    def add_compact_cover(self, company_name, industry, sub_industry):
        """Add a super compact header instead of full cover page"""
        # Simple header bar
        self.set_fill_color(*self.primary_color)
        self.rect(0, 0, 210, 15, 'F')
        
        # Title in header
        self.set_font('Arial', 'B', 12)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 4)
        self.cell(190, 8, "CREDIT RISK REPORT", 0, 1, 'L')
        
        # Company name right below header - no special box
        self.set_font('Arial', 'B', 14)
        self.set_text_color(*self.primary_color)
        self.set_xy(10, 17)
        self.cell(180, 8, company_name.upper(), 0, 1, 'L')
        
        # Industry information inline with smaller font
        if industry:
            self.set_font('Arial', '', 8)
            self.set_xy(10, 24)
            industry_text = f"Industry: {industry}"
            if sub_industry:
                industry_text += f" | {sub_industry}"
            date_text = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
            self.cell(95, 4, industry_text, 0, 0, 'L')
            self.cell(95, 4, date_text, 0, 1, 'R')
        
        # Set position for first section
        self.set_y(30)
        
    def add_section_title(self, title, with_accent=True):
        """Add a very minimal section title"""
        # No background, just text with bottom line
        y_pos = self.get_y()
        
        # Add title 
        self.set_font('Arial', 'B', 10)
        self.set_text_color(*self.primary_color)
        self.set_xy(10, y_pos)
        self.cell(180, 6, title.upper(), 0, 1, 'L')
        
        # Simple line under the title
        if with_accent:
            self.set_draw_color(*self.primary_color)
            self.set_line_width(0.2)
            self.line(10, y_pos + 6, 190, y_pos + 6)
        
        # Minimal space after title
        self.ln(2)
        
    def add_data_table(self, data_dict):
        """Add an extremely compact data table"""
        if not data_dict:
            return
            
        # Very compact column widths
        key_width = 60
        value_width = 120
        
        # No special header styling, just bold text
        self.set_font('Arial', 'B', 8)
        self.set_text_color(*self.primary_color)
        self.set_xy(10, self.get_y() + 2)
        self.cell(key_width, 5, "METRIC", 0, 0, 'L')
        self.cell(value_width, 5, "VALUE", 0, 1, 'L')
        
        # Table content - ultra compact
        self.set_font('Arial', '', 8)
        self.set_text_color(60, 60, 60)
        
        # Process data items in pairs if possible to save vertical space
        items = list(data_dict.items())
        row_y = self.get_y()
        
        for i in range(0, len(items), 2):
            # First item in pair
            key, value = items[i]
            self.set_xy(10, row_y)
            self.set_font('Arial', 'B', 8)
            self.set_text_color(*self.primary_color)
            self.cell(60, 5, key, 0, 0, 'L')
            
            self.set_font('Arial', '', 8)
            self.set_text_color(60, 60, 60)
            self.cell(30, 5, str(value), 0, 0, 'L')
            
            # Second item in pair if available
            if i + 1 < len(items):
                key2, value2 = items[i + 1]
                self.set_xy(110, row_y)
                self.set_font('Arial', 'B', 8)
                self.set_text_color(*self.primary_color)
                self.cell(40, 5, key2, 0, 0, 'L')
                
                self.set_font('Arial', '', 8)
                self.set_text_color(60, 60, 60)
                self.cell(40, 5, str(value2), 0, 0, 'L')
            
            row_y += 5
        
        # Set current position after table
        self.set_y(row_y + 1)
    
    def add_text_block(self, text, is_recommendation=False):
        """Add a very compact text block"""
        if not text:
            return
            
        # Trim excess whitespace for cleaner layout
        text = text.strip()
        
        # Use special minimal styling for recommendations
        if is_recommendation:
            y_start = self.get_y()
            
            # Title for recommendations
            self.set_font('Arial', 'B', 9)
            self.set_text_color(*self.primary_color)
            self.set_xy(10, y_start)
            self.cell(180, 5, "RECOMMENDATIONS:", 0, 1, 'L')
            
            # Add the recommendation text - smaller font
            self.set_font('Arial', '', 8)
            self.set_text_color(60, 60, 60)
            self.set_xy(10, y_start + 5)
            self.multi_cell(180, 4, text, 0, 'L')
            
            # Set position after text
            self.ln(1)
        else:
            # Regular text block with minimal styling
            self.set_font('Arial', '', 8)
            self.set_text_color(60, 60, 60)
            self.set_x(10)
            self.multi_cell(180, 4, text, 0, 'L')
            self.ln(1)
    
    def add_visualization(self, image_data, title=None, max_height=80):
        """Add a very minimal visualization with no decorative elements"""
        if not image_data:
            return
            
        # Calculate dimensions
        y_start = self.get_y() + 1
        max_width = 180
        # Use the parameter instead of hardcoded value
        max_height = max_height  # Adjustable height
        
        # Add title if provided
        if title:
            self.set_font('Arial', 'B', 9)
            self.set_text_color(*self.primary_color)
            self.set_xy(10, y_start)
            self.cell(180, 5, title, 0, 1, 'L')
            y_start += 5
            
        # Process the image based on type
        if isinstance(image_data, str) and image_data.startswith('data:image'):
            # Extract the base64 data
            img_type, data = image_data.split(';base64,')
            img_ext = img_type.split('/')[-1]
            
            # Create temporary file for the image
            temp_img = f"/tmp/temp_viz_{datetime.now().strftime('%Y%m%d%H%M%S')}.{img_ext}"
            with open(temp_img, 'wb') as f:
                f.write(base64.b64decode(data))
            
            # Add the image to the PDF - tighter fit
            self.image(temp_img, x=15, y=y_start, w=max_width - 20, h=max_height)
            
            # Remove the temporary file
            if os.path.exists(temp_img):
                os.remove(temp_img)
        else:
            # If it's a file path
            self.image(image_data, x=15, y=y_start, w=max_width - 20, h=max_height)
            
        # Move down minimally after the visualization
        self.set_y(y_start + max_height + 2)
    
    def rounded_rect(self, x, y, w, h, r, style='', corners='1234'):
        """Draw a rounded rectangle with customizable corners"""
        k = 0.4  # Approximation constant for circular arc
        
        if style == 'F':
            op = 'f'  # Fill
        elif style == 'FD' or style == 'DF':
            op = 'b'  # Fill and border
        else:
            op = 's'  # Border only
            
        # Start point
        self._out(f'{x + r} {y} m')
        
        # Each corner can be selectively rounded based on the corners parameter
        # Top right corner
        if '1' in corners:
            self._out(f'{x + w - r} {y} l')  # Line to start of curve
            self._curve(x + w - r * k, y, x + w, y + r * k, x + w, y + r)
        else:
            self._out(f'{x + w} {y} l')  # Straight corner
            
        # Bottom right corner
        if '2' in corners:
            self._out(f'{x + w} {y + h - r} l')  # Line to start of curve
            self._curve(x + w, y + h - r * k, x + w - r * k, y + h, x + w - r, y + h)
        else:
            self._out(f'{x + w} {y + h} l')  # Straight corner
            
        # Bottom left corner
        if '3' in corners:
            self._out(f'{x + r} {y + h} l')  # Line to start of curve
            self._curve(x + r * k, y + h, x, y + h - r * k, x, y + h - r)
        else:
            self._out(f'{x} {y + h} l')  # Straight corner
            
        # Top left corner
        if '4' in corners:
            self._out(f'{x} {y + r} l')  # Line to start of curve
            self._curve(x, y + r * k, x + r * k, y, x + r, y)
        else:
            self._out(f'{x} {y} l')  # Straight corner
            
        self._out(op)
    
    def _curve(self, x1, y1, x2, y2, x3, y3):
        """Private method: Add a cubic Bezier curve to the document"""
        self._out(f'{x1} {y1} {x2} {y2} {x3} {y3} c')
        
    def ellipse(self, x, y, rx, ry, style=''):
        """Draw an ellipse at (x,y) with radii (rx,ry)"""
        # Style can be 'F' (filled), empty (outlined), or 'FD' (filled and outlined)
        if style == 'F':
            op = 'f'
        elif style == 'FD' or style == 'DF':
            op = 'b'
        else:
            op = 's'
            
        # Approximation of ellipse using cubic Bezier curves
        lx = 4/3 * (math.sqrt(2) - 1) * rx
        ly = 4/3 * (math.sqrt(2) - 1) * ry
        
        # Start point
        self._out(f'{x + rx} {y} m')
        
        # Top right curve
        self._curve(x + rx, y - ly, x + lx, y - ry, x, y - ry)
        
        # Top left curve
        self._curve(x - lx, y - ry, x - rx, y - ly, x - rx, y)
        
        # Bottom left curve
        self._curve(x - rx, y + ly, x - lx, y + ry, x, y + ry)
        
        # Bottom right curve
        self._curve(x + lx, y + ry, x + rx, y + ly, x + rx, y)
        
        self._out(op)

class ReportGenerator:
    def __init__(self, template_dir='templates', output_dir='reports'):
        self.template_dir = template_dir
        self.output_dir = output_dir
        self.env = Environment(loader=FileSystemLoader(template_dir))
        
        # Live Oak Bank color palette
        self.colors = {
            'primary': '#1E1847',
            'primary_light': '#e8e7f0',
            'accent': '#2DC77A',
            'accent_light': '#e9f9f2',
            'positive': '#2DC77A',
            'negative': '#ef4444',
            'neutral': '#64748b',
            'high_risk': '#ef4444',
            'medium_risk': '#f59e0b',
            'low_risk': '#2DC77A',
            'background': '#f8fafc',
            'text': '#1e293b'
        }
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def generate_report(self, company_name, data):
        """
        Generate a report for the given company using the provided data
        
        Args:
            company_name (str): Name of the company
            data (dict): Dictionary containing the data for the report
            
        Returns:
            str: Path to the generated report file
        """
        try:
            # Prepare the template context
            context = {
                'company_name': company_name,
                'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'year': datetime.now().year,
                'sections': self._prepare_sections(data)
            }
            
            # Render the template
            template = self.env.get_template('report_template.html')
            html_content = template.render(**context)
            
            # Generate unique filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{company_name.replace(' ', '_')}_{timestamp}.html"
            filepath = os.path.join(self.output_dir, filename)
            
            # Save the report
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            return filepath
        except Exception as e:
            print(f"Error generating report: {str(e)}")
            raise

    def _prepare_sections(self, report_data):
        """Prepare sections for the report with enhanced visualizations"""
        sections = []
        
        # Add each section to the report
        for section in report_data['sections']:
            section_data = {
                'title': section['title'],
                'data': section['data']
            }
            
            # Generate visualizations based on section type
            if 'risk_profile' in section.get('section_type', ''):
                if isinstance(section['data'], dict):
                    section_data['visualization'] = self._create_risk_profile_visualization(section['data'])
            
            elif 'financial_metrics' in section.get('section_type', ''):
                if isinstance(section['data'], dict):
                    section_data['visualization'] = self._create_financial_metrics_visualization(section['data'])
            
            # Only add visualizations if explicitly requested
            elif 'visualizations' in section and section.get('include_visualizations', False):
                section_data['visualizations'] = []
                for viz_data in section['visualizations']:
                    viz = self._create_visualization(viz_data)
                    if viz:
                        section_data['visualizations'].append(viz)
            
            sections.append(section_data)
        
        return sections

    def _create_risk_profile_visualization(self, data):
        """Create a visualization specifically for risk profiles"""
        # Extract risk data
        risk_level = data.get('Risk Level', 'Medium')
        pd = float(data.get('Probability of Default', '0.0%').strip('%')) / 100
        lgd = float(data.get('Loss Given Default', '0.0%').strip('%')) / 100
        expected_loss = data.get('Expected Loss', '$0.00').strip('$').replace(',', '')
        
        try:
            expected_loss = float(expected_loss)
        except:
            expected_loss = 0.0
        
        # Create figure with multiple traces
        fig = go.Figure()
        
        # Risk gauge
        fig.add_trace(go.Indicator(
            mode = "gauge+number+delta",
            value = pd * 100,
            domain = {'x': [0, 1], 'y': [0.5, 1]},
            title = {'text': "Probability of Default (%)", 'font': {'size': 16, 'color': self.colors['text']}},
            gauge = {
                'axis': {'range': [None, 10], 'tickwidth': 1, 'tickcolor': self.colors['text']},
                'bar': {'color': self.colors['primary']},
                'bgcolor': self.colors['background'],
                'borderwidth': 2,
                'bordercolor': self.colors['primary_light'],
                'steps': [
                    {'range': [0, 1], 'color': self.colors['low_risk']},
                    {'range': [1, 5], 'color': self.colors['medium_risk']},
                    {'range': [5, 10], 'color': self.colors['high_risk']}
                ],
                'threshold': {
                    'line': {'color': self.colors['primary'], 'width': 4},
                    'thickness': 0.75,
                    'value': pd * 100
                }
            }
        ))
        
        # LGD and Expected Loss Bar Chart
        fig.add_trace(go.Bar(
            x=['Loss Given Default (%)', 'Expected Loss ($ thousands)'],
            y=[lgd * 100, expected_loss / 1000],
            text=[f"{lgd*100:.1f}%", f"${expected_loss/1000:.1f}K"],
            textposition='auto',
            marker_color=[self.colors['primary'], self.colors['accent']],
            hoverinfo='y+text'
        ))
        
        # Customize layout
        fig.update_layout(
            title={
                'text': 'Risk Assessment Overview',
                'y':0.95,
                'x':0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': {'size': 20, 'color': self.colors['primary']}
            },
            paper_bgcolor=self.colors['background'],
            plot_bgcolor=self.colors['background'],
            font={'color': self.colors['text'], 'family': 'Inter, Arial, sans-serif'},
            margin=dict(l=20, r=20, t=80, b=20),
            height=500,
            grid={'rows': 2, 'columns': 1, 'pattern': 'independent'},
            showlegend=False
        )
        
        # Convert to image
        img_bytes = BytesIO()
        fig.write_image(img_bytes, format='png', scale=2, engine='kaleido')
        img_bytes.seek(0)
        
        # Convert to base64
        img_str = base64.b64encode(img_bytes.getvalue()).decode()
        return f"data:image/png;base64,{img_str}"

    def _create_financial_metrics_visualization(self, data):
        """Create a visualization specifically for financial metrics"""
        # Extract financial data
        metrics = []
        values = []
        colors = []
        
        if 'Current Ratio' in data:
            metrics.append('Current Ratio')
            current_ratio = float(data.get('Current Ratio', '0.00'))
            values.append(current_ratio)
            colors.append(self.colors['primary'] if current_ratio >= 1.5 else 
                          self.colors['medium_risk'] if current_ratio >= 1.0 else 
                          self.colors['high_risk'])
        
        if 'ROA' in data:
            metrics.append('ROA (%)')
            roa = float(data.get('ROA', '0.00%').strip('%'))
            values.append(roa)
            colors.append(self.colors['positive'] if roa > 0 else self.colors['negative'])
        
        if 'ROE' in data:
            metrics.append('ROE (%)')
            roe = float(data.get('ROE', '0.00%').strip('%'))
            values.append(roe)
            colors.append(self.colors['positive'] if roe > 0 else self.colors['negative'])
        
        if 'Leverage Ratio' in data:
            metrics.append('Leverage Ratio')
            leverage = float(data.get('Leverage Ratio', '0.00'))
            values.append(leverage)
            colors.append(self.colors['primary'] if leverage <= 2.0 else 
                          self.colors['medium_risk'] if leverage <= 3.0 else 
                          self.colors['high_risk'])
        
        if 'Debt Service Coverage Ratio' in data:
            metrics.append('DSCR')
            dscr = float(data.get('Debt Service Coverage Ratio', '0.00'))
            values.append(dscr)
            colors.append(self.colors['positive'] if dscr >= 1.25 else 
                          self.colors['medium_risk'] if dscr >= 1.0 else 
                          self.colors['negative'])
        
        # Create bar chart
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            x=metrics,
            y=values,
            text=[f"{v:.2f}" for v in values],
            textposition='auto',
            marker_color=colors,
            hoverinfo='y+text'
        ))
        
        # Add reference lines for key thresholds
        if 'Current Ratio' in metrics:
            idx = metrics.index('Current Ratio')
            fig.add_shape(type="line",
                x0=idx-0.4, y0=1.5, x1=idx+0.4, y1=1.5,
                line=dict(color=self.colors['low_risk'], width=2, dash="dash"))
        
        if 'DSCR' in metrics:
            idx = metrics.index('DSCR')
            fig.add_shape(type="line",
                x0=idx-0.4, y0=1.25, x1=idx+0.4, y1=1.25,
                line=dict(color=self.colors['low_risk'], width=2, dash="dash"))
        
        # Customize layout
        fig.update_layout(
            title={
                'text': 'Financial Performance Metrics',
                'y':0.95,
                'x':0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': {'size': 20, 'color': self.colors['primary']}
            },
            paper_bgcolor=self.colors['background'],
            plot_bgcolor=self.colors['background'],
            font={'color': self.colors['text'], 'family': 'Inter, Arial, sans-serif'},
            margin=dict(l=20, r=20, t=80, b=40),
            height=400,
            yaxis=dict(
                title=dict(
                    text="Value",
                    font=dict(size=14)
                ),
                gridcolor='rgba(230, 233, 237, 0.7)'
            ),
            xaxis=dict(
                title=dict(
                    text="Metric",
                    font=dict(size=14)
                )
            ),
            showlegend=False
        )
        
        # Convert to image
        img_bytes = BytesIO()
        fig.write_image(img_bytes, format='png', scale=2, engine='kaleido')
        img_bytes.seek(0)
        
        # Convert to base64
        img_str = base64.b64encode(img_bytes.getvalue()).decode()
        return f"data:image/png;base64,{img_str}"

    def _create_visualization(self, viz_data):
        """
        Create a visualization using plotly and return it as a base64 encoded image
        
        Args:
            viz_data (dict): Visualization data and configuration
            
        Returns:
            str: Base64 encoded image
        """
        fig = go.Figure()
        
        # Set theme colors
        template = {
            'layout': {
                'colorway': [self.colors['primary'], self.colors['accent'], 
                            '#4338ca', '#0891b2', '#a855f7', 
                            '#ec4899', '#f59e0b', '#84cc16'],
                'paper_bgcolor': self.colors['background'],
                'plot_bgcolor': self.colors['background'],
                'font': {'color': self.colors['text'], 'family': 'Inter, Arial, sans-serif'},
                'title': {'font': {'color': self.colors['primary'], 'size': 20}},
                'legend': {'font': {'color': self.colors['text']}},
                'xaxis': {
                    'gridcolor': 'rgba(230, 233, 237, 0.7)',
                    'title': {'font': {'color': self.colors['text'], 'size': 14}},
                    'tickfont': {'color': self.colors['text']}
                },
                'yaxis': {
                    'gridcolor': 'rgba(230, 233, 237, 0.7)',
                    'title': {'font': {'color': self.colors['text'], 'size': 14}},
                    'tickfont': {'color': self.colors['text']}
                }
            }
        }
        
        if viz_data['type'] == 'line':
            fig.add_trace(go.Scatter(
                x=viz_data['x'],
                y=viz_data['y'],
                mode='lines+markers',
                name=viz_data.get('name', ''),
                line=dict(color=self.colors['primary'], width=3),
                marker=dict(color=self.colors['accent'], size=8, line=dict(width=2, color=self.colors['primary']))
            ))
            
        elif viz_data['type'] == 'bar':
            fig.add_trace(go.Bar(
                x=viz_data['x'],
                y=viz_data['y'],
                name=viz_data.get('name', ''),
                marker_color=self.colors['primary'],
                text=viz_data['y'],
                textposition='auto'
            ))
            
        elif viz_data['type'] == 'pie':
            fig = go.Figure(data=[go.Pie(
                labels=viz_data['labels'],
                values=viz_data['values'],
                textinfo='label+percent',
                insidetextorientation='radial',
                marker=dict(
                    colors=[self.colors['primary'], self.colors['accent'], 
                           '#4338ca', '#0891b2', '#a855f7']
                )
            )])
            
        elif viz_data['type'] == 'scatter':
            fig.add_trace(go.Scatter(
                x=viz_data['x'],
                y=viz_data['y'],
                mode='markers',
                name=viz_data.get('name', ''),
                marker=dict(
                    color=self.colors['primary'],
                    size=10,
                    opacity=0.7,
                    line=dict(width=1, color=self.colors['accent'])
                )
            ))
        
        # Enhanced layout with Live Oak Bank theme
        fig.update_layout(
            title=viz_data['title'],
            xaxis_title=viz_data.get('xaxis_title', ''),
            yaxis_title=viz_data.get('yaxis_title', ''),
            template=template,
            margin=dict(l=20, r=20, t=60, b=40),
            height=400
        )
        
        # Apply light gridlines
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(230, 233, 237, 0.7)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(230, 233, 237, 0.7)')
        
        # Convert to image with higher resolution
        img_bytes = BytesIO()
        fig.write_image(img_bytes, format='png', scale=2, engine='kaleido')
        img_bytes.seek(0)
        
        # Convert to base64
        img_str = base64.b64encode(img_bytes.getvalue()).decode()
        return f"data:image/png;base64,{img_str}"

    def _generate_visualizations(self, data: Dict, report_dir: str) -> Dict[str, str]:
        """Generate requested visualizations"""
        viz_paths = {}
        
        for viz in data.get('visualizations', []):
            if viz == 'risk_distribution':
                path = self._create_risk_distribution_chart(data, report_dir)
                viz_paths['risk_distribution'] = path
            elif viz == 'financial_trends':
                path = self._create_financial_trends_chart(data, report_dir)
                viz_paths['financial_trends'] = path
            elif viz == 'industry_comparison':
                path = self._create_industry_comparison_chart(data, report_dir)
                viz_paths['industry_comparison'] = path
            elif viz == 'news_sentiment':
                path = self._create_news_sentiment_chart(data, report_dir)
                viz_paths['news_sentiment'] = path
        
        return viz_paths
    
    def _generate_pdf_report(self, report_data, template_context, output_dir):
        """Generate a single-page compact PDF report with modern design elements and professional styling"""
        try:
            from fpdf import FPDF, HTMLMixin
            import os
            from datetime import datetime
            
            # Ensure output directory is absolute path
            output_dir = os.path.abspath(output_dir)
            
            # Make sure output directory exists
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                print(f"Created output directory: {output_dir}")
            
            print(f"Generating PDF report in directory: {output_dir}")
                
            # Extract company info
            company_name = report_data['company_name']
            industry = ''
            sub_industry = ''
            for section in report_data.get('sections', []):
                if section.get('section_type') == 'company_info' and isinstance(section.get('data'), dict):
                    industry = section['data'].get('Industry', '')
                    sub_industry = section['data'].get('Sub-Industry', '')
                    break
            
            # Create our premium PDF with ultra-tight margins
            pdf = PremiumPDF(orientation='P', unit='mm', format='A4')
            pdf.set_margins(5, 5, 5)  # Ultra tight margins
            pdf.set_auto_page_break(False)  # No auto page breaks - force single page
            
            # Start with a single page
            pdf.add_page()
            
            # Add compact header (not a full cover)
            pdf.add_compact_cover(company_name, industry, sub_industry)
            
            # Define section order by type
            section_order = ['company_info', 'risk_profile', 'financial_metrics', 'dscr_analysis']
            
            # Map sections to their types
            sections_by_type = {}
            
            # Group sections by type for ordered processing
            for section in report_data.get('sections', []):
                section_type = section.get('section_type', '').lower()
                if section_type:
                    for order_type in section_order:
                        if order_type in section_type:
                            sections_by_type[order_type] = section
                            break
            
            # Current Y position for content
            current_y = 32  # Starting position after header
            
            # Add sections in specified order
            for section_type in section_order:
                if section_type in sections_by_type:
                    section = sections_by_type[section_type]
                    
                    # Set y position
                    pdf.set_y(current_y)
                    
                    # Add section title (minimal styling)
                    section_title = section.get('title', section_type.replace('_', ' ').title())
                    pdf.add_section_title(section_title)
                    
                    # Add section content based on data type
                    section_data = section.get('data', {})
                    
                    if isinstance(section_data, dict):
                        pdf.add_data_table(section_data)
                    else:
                        is_recommendation = 'recommendation' in section_type
                        # Truncate text if too long to fit on one page
                        text = str(section_data)
                        if len(text) > 300:
                            text = text[:300] + "..."
                        pdf.add_text_block(text, is_recommendation)
                    
                    # Add visualization if available and we have space
                    if 'visualization' in section and section['visualization']:
                        # Determine visualization height based on remaining space and priority
                        remaining_sections = sum(1 for st in section_order[section_order.index(section_type)+1:] 
                                                if st in sections_by_type)
                        # Less height for visualizations if more sections follow
                        max_height = 80 if remaining_sections == 0 else 50
                        
                        # Only add visualization if we're not near bottom of page
                        current_pos = pdf.get_y()
                        if current_pos < 240:
                            pdf.add_visualization(section['visualization'], None, max_height=max_height)
                    
                    # Update position for next section
                    current_y = pdf.get_y() + 3
            
            # Add minimal footer
            pdf.set_font('Arial', 'I', 7)
            pdf.set_text_color(150, 150, 150)
            pdf.set_xy(10, 290)
            pdf.cell(190, 5, f"Generated: {datetime.now().strftime('%Y-%m-%d')} | Confidential: For authorized use only", 0, 0, 'C')
            
            # Generate unique filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # Create safe filename
            safe_company_name = ''.join(c if c.isalnum() else '_' for c in company_name)
            filename = f"{safe_company_name}_{timestamp}.pdf"
            filepath = os.path.join(output_dir, filename)
            
            print(f"About to save PDF to: {filepath}")
            
            # Save the PDF
            pdf.output(filepath)
            
            # Verify file was created
            if os.path.exists(filepath):
                file_size = os.path.getsize(filepath)
                print(f"PDF created successfully at {filepath}, size: {file_size} bytes")
            else:
                print(f"WARNING: PDF file not found at {filepath} after generation attempt")
                
            return filepath
            
        except Exception as e:
            print(f"Error generating PDF report: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _generate_excel_report(self, report_data, template_context, output_dir):
        """Generate Excel report"""
        try:
            # Create a new Excel workbook
            wb = Workbook()
            
            # Add company info sheet
            ws = wb.active
            ws.title = "Company Information"
            
            # Add company name
            ws['A1'] = "Company Name"
            ws['B1'] = report_data['company_name']
            
            # Add report generation date
            ws['A2'] = "Generated Date"
            ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Add sections
            row = 4
            for section in report_data['sections']:
                # Add section title with formatting
                ws[f'A{row}'] = section['title']
                ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=14)
                row += 1
                
                if isinstance(section['data'], dict):
                    # Add headers with formatting
                    for col, header in enumerate(section['data'].keys(), 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    row += 1
                    
                    # Add data
                    for col, value in enumerate(section['data'].values(), 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        # Set text wrapping for cells with long content
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    row += 1
                else:
                    # Add text content with wrapping
                    cell = ws[f'A{row}']
                    cell.value = section['data']
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    row += 1
                
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col)
                
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Generate unique filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{report_data['company_name'].replace(' ', '_')}_{timestamp}.xlsx"
            output_path = os.path.join(output_dir, filename)
            
            # Save workbook
            wb.save(output_path)
            
            return output_path
            
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _generate_html_report(self, report_data, template_context, output_dir):
        """Generate HTML report"""
        try:
            print(f"Starting HTML report generation for {report_data.get('company_name')}")
            print(f"Template context: {template_context}")
            print(f"Output directory: {output_dir}")
            
            # Prepare template context
            context = self._prepare_template_context(report_data, template_context)
            print(f"Prepared context: {context}")
            
            # Get the template
            template = self.env.get_template('report_template.html')
            print("Successfully loaded template")
            
            # Generate HTML content
            html_content = template.render(**context)
            print("Successfully rendered HTML content")
            
            # Generate unique filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{report_data['company_name'].replace(' ', '_')}_{timestamp}.html"
            output_path = os.path.join(output_dir, filename)
            print(f"Output path: {output_path}")
            
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            print(f"Ensured output directory exists: {output_dir}")
            
            # Save HTML file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            print(f"Successfully wrote HTML file to {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"Error generating HTML report: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _prepare_template_context(self, report_data, template_context):
        """Prepare context for template rendering"""
        context = {
            'company_name': report_data['company_name'],
            'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'sections': report_data['sections']
        }
        
        # Add any additional context from template_context
        if template_context:
            context.update(template_context)
        
        return context

    def _add_section_to_pdf(self, pdf: FPDF, section: str, data: Dict, viz_paths: Dict[str, str]):
        """Add a section to the PDF report"""
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, section.replace('_', ' ').title(), ln=True)
        pdf.ln(5)
        
        # Add section content
        pdf.set_font('Arial', '', 12)
        if section == 'company_info':
            self._add_company_info_to_pdf(pdf, data)
        elif section == 'risk_profile':
            self._add_risk_profile_to_pdf(pdf, data)
        elif section == 'financial_metrics':
            self._add_financial_metrics_to_pdf(pdf, data)
        # Add more section handlers as needed
        
        # Add visualization if available
        if section in viz_paths:
            pdf.image(viz_paths[section], x=10, y=None, w=190)
            pdf.ln(10)
    
    def _add_section_to_excel(self, worksheet, section: str, data: Dict):
        """Add a section to the Excel report"""
        if section == 'company_info':
            self._add_company_info_to_excel(worksheet, data)
        elif section == 'risk_profile':
            self._add_risk_profile_to_excel(worksheet, data)
        elif section == 'financial_metrics':
            self._add_financial_metrics_to_excel(worksheet, data)
        # Add more section handlers as needed
    
    def _create_risk_distribution_chart(self, data: Dict, report_dir: str) -> str:
        """Create risk distribution chart with modern styling"""
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib import cm
            
            # Create figure with higher DPI for better quality
            plt.figure(figsize=(10, 6), dpi=120)
            
            # Get risk distribution data
            risk_data = data.get('risk_distribution', {
                'High Risk': 15,
                'Medium Risk': 45,
                'Low Risk': 40
            })
            
            # Create colors for each risk level with modern palette
            colors = ['#EF4444', '#F59E0B', '#2DC77A']  # Red, Orange, Green
            
            # Create explode to emphasize high risk
            explode = (0.1, 0.05, 0)
            
            # Create donut chart
            wedges, texts, autotexts = plt.pie(
                list(risk_data.values()),
                labels=list(risk_data.keys()),
                autopct='%1.1f%%',
                startangle=90,
                explode=explode,
                colors=colors,
                wedgeprops={'width': 0.5, 'edgecolor': 'white', 'linewidth': 2},
                textprops={'fontsize': 12, 'fontweight': 'bold', 'color': '#1E1847'}
            )
            
            # Equal aspect ratio ensures that pie is drawn as a circle
            plt.axis('equal')
            
            # Style the percentage texts
            for autotext in autotexts:
                autotext.set_fontsize(10)
                autotext.set_fontweight('bold')
                autotext.set_color('white')
            
            # Add title with styling
            plt.title('Risk Distribution', fontsize=16, fontweight='bold', color='#1E1847', pad=20)
            
            # Add subtle grid
            plt.grid(False)
            
            # Add company name as subtitle
            if 'company_name' in data:
                plt.figtext(0.5, 0.02, f"Company: {data['company_name']}", ha="center", fontsize=10, color='#64748b')
            
            # Set background color to white
            plt.gca().set_facecolor('white')
            fig = plt.gcf()
            fig.patch.set_facecolor('white')
            
            # Add a circle at the center to create a donut chart
            circle = plt.Circle((0, 0), 0.25, fc='white')
            plt.gca().add_artist(circle)
            
            # Save chart
            output_path = os.path.join(report_dir, 'risk_distribution.png')
            plt.savefig(output_path, bbox_inches='tight', facecolor='white', dpi=120)
            plt.close()
            return output_path
            
        except Exception as e:
            print(f"Error creating risk distribution chart: {str(e)}")
            import traceback
            traceback.print_exc()
            # Return a placeholder path
            return ""
    
    def _create_financial_trends_chart(self, data: Dict, report_dir: str) -> str:
        """Create financial trends chart with modern styling"""
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            import matplotlib.dates as mdates
            from matplotlib.ticker import FuncFormatter
            
            # Create figure with higher DPI for better quality
            fig, ax = plt.subplots(figsize=(11, 6), dpi=120)
            
            # Sample financial data over time
            financial_metrics = data.get('financial_trends', {
                'Revenue': [10.2, 12.5, 15.1, 16.8, 18.4, 20.1],
                'Profit Margin': [0.12, 0.14, 0.15, 0.13, 0.16, 0.17],
                'Cash Flow': [2.1, 2.4, 3.0, 2.8, 3.2, 3.5],
                'Debt Ratio': [0.45, 0.42, 0.38, 0.35, 0.32, 0.30]
            })
            
            # Get time periods (last 6 quarters by default)
            periods = data.get('time_periods', [f'Q{i} {2023 if i > 2 else 2024}' for i in range(1, 7)])
            
            # Create modern color palette
            colors = {
                'Revenue': '#2DC77A',         # Green
                'Profit Margin': '#1E88E5',   # Blue
                'Cash Flow': '#9C27B0',       # Purple
                'Debt Ratio': '#E53935'       # Red
            }
            
            # Line styles
            line_styles = {
                'Revenue': '-',
                'Profit Margin': '--',
                'Cash Flow': '-.',
                'Debt Ratio': ':'
            }
            
            # Settings for primary axis (left)
            ax.set_xlabel('Quarter', fontsize=12, color='#1E1847')
            ax.set_ylabel('USD (Millions)', fontsize=12, color='#1E1847')
            
            # Create a secondary axis for percentages
            ax2 = ax.twinx()
            ax2.set_ylabel('Percentage', fontsize=12, color='#1E1847')
            
            # Track which metrics are percentages
            percentage_metrics = ['Profit Margin', 'Debt Ratio']
            
            # Plot each metric
            for metric, values in financial_metrics.items():
                if metric in percentage_metrics:
                    # Plot on secondary axis (percentages)
                    ax2.plot(periods, values, marker='o', linestyle=line_styles.get(metric, '-'), 
                            linewidth=2.5, color=colors.get(metric, '#000000'), 
                            label=f"{metric} (%)")
                else:
                    # Plot on primary axis (dollar values)
                    ax.plot(periods, values, marker='s', linestyle=line_styles.get(metric, '-'), 
                           linewidth=2.5, color=colors.get(metric, '#000000'), 
                           label=metric)
            
            # Format the y-axis ticks for primary axis (dollar values)
            def millions(x, pos):
                return f'${x:.1f}M'
            
            ax.yaxis.set_major_formatter(FuncFormatter(millions))
            
            # Format the percentage values on secondary axis
            def percentage(x, pos):
                return f'{x*100:.0f}%'
                
            ax2.yaxis.set_major_formatter(FuncFormatter(percentage))
            
            # Set grid style
            ax.grid(True, linestyle='--', alpha=0.3)
            
            # Combine legends from both axes
            lines1, labels1 = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            
            # Add legend with custom styling
            ax.legend(lines1 + lines2, labels1 + labels2, loc='upper center', 
                     bbox_to_anchor=(0.5, -0.15), ncol=4, frameon=True, 
                     facecolor='white', edgecolor='#E6E9ED', fontsize=10)
            
            # Add title
            plt.title('Financial Performance Trends', fontsize=16, fontweight='bold', color='#1E1847', pad=20)
            
            # Style the ticks and spines
            for ax_i in [ax, ax2]:
                ax_i.tick_params(axis='both', labelsize=10, colors='#64748b')
                for spine in ax_i.spines.values():
                    spine.set_color('#E6E9ED')
            
            # Annotate growth rates for important metrics
            for i, metric in enumerate(['Revenue', 'Profit Margin']):
                if metric in financial_metrics:
                    values = financial_metrics[metric]
                    # Calculate growth from first to last period
                    growth = ((values[-1] - values[0]) / values[0]) * 100
                    color = '#2DC77A' if growth >= 0 else '#EF4444'
                    
                    # Choose axis for annotation
                    current_ax = ax2 if metric in percentage_metrics else ax
                    
                    # Add annotation arrow and text
                    current_ax.annotate(
                        f"{growth:.1f}% growth",
                        xy=(periods[-1], values[-1]),
                        xytext=(15, 15 * (i+1)),
                        textcoords="offset points",
                        arrowprops=dict(arrowstyle="->", color=color),
                        fontsize=9,
                        fontweight='bold',
                        color=color
                    )
            
            # Adjust layout to make room for the legend
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.25)
            
            # Set background color
            fig.patch.set_facecolor('white')
            ax.set_facecolor('#FAFAFA')
            
            # Save figure
            output_path = os.path.join(report_dir, 'financial_trends.png')
            plt.savefig(output_path, bbox_inches='tight', facecolor='white', dpi=120)
            plt.close()
            return output_path
            
        except Exception as e:
            print(f"Error creating financial trends chart: {str(e)}")
            import traceback
            traceback.print_exc()
            return ""
    
    def _create_industry_comparison_chart(self, data: Dict, report_dir: str) -> str:
        """Create industry comparison chart with modern styling"""
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib.ticker import FuncFormatter
            
            # Create figure with higher DPI for better quality
            fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
            
            # Sample industry comparison data (company vs. industry average)
            metrics = data.get('industry_metrics', [
                'Revenue Growth', 'Profit Margin', 'ROA', 'ROE', 'Current Ratio', 'Debt to Equity'
            ])
            
            company_values = data.get('company_values', [15, 12, 8, 14, 2.2, 0.8])
            industry_values = data.get('industry_values', [10, 10, 7, 12, 1.8, 1.2])
            
            # Number of metrics for x-axis
            x = np.arange(len(metrics))
            width = 0.35  # Width of the bars
            
            # Create bars with modern styling
            company_bars = ax.bar(x - width/2, company_values, width, label='Company', 
                                 color='#2DC77A', edgecolor='white', linewidth=1, alpha=0.9)
            industry_bars = ax.bar(x + width/2, industry_values, width, label='Industry Average', 
                                  color='#64748b', edgecolor='white', linewidth=1, alpha=0.7)
            
            # Add value labels on top of bars
            def add_labels(bars):
                for bar in bars:
                    height = bar.get_height()
                    is_percentage = height < 100 and height > 0
                    
                    ax.annotate(f"{height:.1f}{'%' if is_percentage else ''}",
                               xy=(bar.get_x() + bar.get_width() / 2, height),
                               xytext=(0, 3),  # 3 points vertical offset
                               textcoords="offset points",
                               ha='center', va='bottom',
                               fontsize=9, fontweight='bold')
            
            add_labels(company_bars)
            add_labels(industry_bars)
            
            # Customize axis
            ax.set_ylabel('Value', fontsize=12, color='#1E1847')
            ax.set_title('Company vs. Industry Average', fontsize=16, fontweight='bold', color='#1E1847', pad=20)
            ax.set_xticks(x)
            ax.set_xticklabels(metrics, rotation=25, ha='right', fontsize=10)
            
            # Add grid lines for better readability
            ax.grid(axis='y', linestyle='--', alpha=0.3)
            
            # Add legend with custom styling
            ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), 
                     ncol=2, frameon=True, facecolor='white', 
                     edgecolor='#E6E9ED', fontsize=10)
            
            # Style the ticks and spines
            ax.tick_params(axis='both', labelsize=10, colors='#64748b')
            for spine in ax.spines.values():
                spine.set_color('#E6E9ED')
            
            # Highlight areas where company outperforms industry
            for i in range(len(metrics)):
                if company_values[i] > industry_values[i]:
                    # Consider if higher is better for this metric
                    if metrics[i] != 'Debt to Equity':  # For debt, lower is better
                        ax.annotate('↑', xy=(x[i], max(company_values[i], industry_values[i]) + 0.5),
                                  xytext=(0, 5), textcoords='offset points',
                                  ha='center', fontsize=14, color='#2DC77A', fontweight='bold')
                elif company_values[i] < industry_values[i]:
                    # Consider if lower is better for this metric
                    if metrics[i] == 'Debt to Equity':  # For debt, lower is better
                        ax.annotate('↑', xy=(x[i], min(company_values[i], industry_values[i]) - 1),
                                  xytext=(0, -15), textcoords='offset points',
                                  ha='center', fontsize=14, color='#2DC77A', fontweight='bold')
            
            # Set background color
            fig.patch.set_facecolor('white')
            ax.set_facecolor('#FAFAFA')
            
            # Adjust layout
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.25)
            
            # Save chart
            output_path = os.path.join(report_dir, 'industry_comparison.png')
            plt.savefig(output_path, bbox_inches='tight', facecolor='white', dpi=120)
            plt.close()
            return output_path
            
        except Exception as e:
            print(f"Error creating industry comparison chart: {str(e)}")
            import traceback
            traceback.print_exc()
            return ""
    
    def _create_news_sentiment_chart(self, data: Dict, report_dir: str) -> str:
        """Create news sentiment chart with modern styling"""
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib.ticker import MaxNLocator
            import matplotlib.dates as mdates
            from datetime import datetime, timedelta
            
            # Create figure with higher DPI for better quality
            fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
            
            # Generate sample dates for the news (last 30 days)
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            # Sample news sentiment data
            dates = data.get('news_dates', [
                end_date - timedelta(days=i) for i in range(30, 0, -5)
            ])
            
            sentiment_scores = data.get('sentiment_scores', [
                0.72, -0.45, 0.2, 0.8, -0.3, 0.15
            ])
            
            # Number of news articles per date
            news_counts = data.get('news_counts', [
                5, 3, 2, 4, 2, 3
            ])
            
            # Create scatter plot for sentiment by date
            scatter = ax.scatter(dates, sentiment_scores, 
                               s=[count * 50 for count in news_counts],  # Size based on number of articles
                               c=[self._get_sentiment_color(score) for score in sentiment_scores],
                               alpha=0.7, edgecolor='white', linewidth=1)
            
            # Add a horizontal line at y=0 to separate positive/negative sentiment
            ax.axhline(y=0, color='#E6E9ED', linestyle='-', linewidth=1)
            
            # Fill areas above/below the sentiment line
            ax.fill_between(dates, sentiment_scores, 0, 
                          where=[score > 0 for score in sentiment_scores],
                          color='#2DC77A', alpha=0.2)
            ax.fill_between(dates, sentiment_scores, 0, 
                          where=[score <= 0 for score in sentiment_scores],
                          color='#EF4444', alpha=0.2)
            
            # Format the x-axis with dates
            date_format = mdates.DateFormatter('%b %d')
            ax.xaxis.set_major_formatter(date_format)
            plt.xticks(rotation=45)
            
            # Format y-axis to show sentiment as percentage
            def sentiment_pct(x, pos):
                return f"{int(x*100)}%"
                
            ax.yaxis.set_major_formatter(FuncFormatter(sentiment_pct))
            
            # Set y-axis limits with padding
            ax.set_ylim(-1.1, 1.1)
            
            # Add labels and title
            ax.set_xlabel('Date', fontsize=12, color='#1E1847')
            ax.set_ylabel('Sentiment Score', fontsize=12, color='#1E1847')
            ax.set_title('News Sentiment Analysis', fontsize=16, fontweight='bold', color='#1E1847', pad=20)
            
            # Add sentiment zones with labels
            ax.annotate('Positive Sentiment', xy=(0.85, 0.85), xycoords='axes fraction',
                      fontsize=10, color='#2DC77A', fontweight='bold')
            ax.annotate('Negative Sentiment', xy=(0.85, 0.15), xycoords='axes fraction',
                      fontsize=10, color='#EF4444', fontweight='bold')
            
            # Add a legend explaining the bubble size
            max_count = max(news_counts)
            min_count = min(news_counts)
            mid_count = (max_count + min_count) // 2
            
            # Create custom legend for bubble sizes
            legend_elements = [
                plt.Line2D([0], [0], marker='o', color='w', label=f'{min_count} Articles',
                          markerfacecolor='#64748b', markersize=np.sqrt(min_count * 5)),
                plt.Line2D([0], [0], marker='o', color='w', label=f'{mid_count} Articles',
                          markerfacecolor='#64748b', markersize=np.sqrt(mid_count * 5)),
                plt.Line2D([0], [0], marker='o', color='w', label=f'{max_count} Articles',
                          markerfacecolor='#64748b', markersize=np.sqrt(max_count * 5))
            ]
            
            ax.legend(handles=legend_elements, loc='upper left', 
                     frameon=True, facecolor='white', edgecolor='#E6E9ED')
            
            # Add grid
            ax.grid(True, linestyle='--', alpha=0.3)
            
            # Style the ticks and spines
            ax.tick_params(axis='both', labelsize=10, colors='#64748b')
            for spine in ax.spines.values():
                spine.set_color('#E6E9ED')
            
            # Set background color
            fig.patch.set_facecolor('white')
            ax.set_facecolor('#FAFAFA')
            
            # Calculate and display average sentiment
            avg_sentiment = sum(sentiment_scores) / len(sentiment_scores)
            sentiment_text = f"Average Sentiment: {avg_sentiment:.2f}"
            ax.annotate(sentiment_text, xy=(0.02, 0.95), xycoords='axes fraction',
                      fontsize=11, fontweight='bold',
                      color='#2DC77A' if avg_sentiment > 0 else '#EF4444')
            
            # Adjust layout
            plt.tight_layout()
            
            # Save the chart
            output_path = os.path.join(report_dir, 'news_sentiment.png')
            plt.savefig(output_path, bbox_inches='tight', facecolor='white', dpi=120)
            plt.close()
            return output_path
            
        except Exception as e:
            print(f"Error creating news sentiment chart: {str(e)}")
            import traceback
            traceback.print_exc()
            return ""
    
    def _get_sentiment_color(self, sentiment_score):
        """Get color based on sentiment score for visualization"""
        if sentiment_score > 0.5:
            return '#10B981'  # Strong positive - green
        elif sentiment_score > 0:
            return '#34D399'  # Mild positive - light green
        elif sentiment_score > -0.5:
            return '#F87171'  # Mild negative - light red
        else:
            return '#EF4444'  # Strong negative - red

    def _get_standard_template(self) -> List[str]:
        """Get standard report template sections"""
        return ['company_info', 'risk_profile', 'financial_metrics']
    
    def _get_executive_template(self) -> List[str]:
        """Get executive report template sections"""
        return ['company_info', 'risk_profile', 'recommendations']
    
    def _get_detailed_template(self) -> List[str]:
        """Get detailed report template sections"""
        return [
            'company_info',
            'risk_profile',
            'financial_metrics',
            'historical_data',
            'industry_comparison',
            'news_analysis',
            'recommendations'
        ]
    
    def _get_html_template(self) -> Template:
        """Get HTML report template"""
        template_str = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Report for {{ company_name }}</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; }
                h1 { color: #333; }
                .section { margin: 20px 0; }
                .visualization { margin: 20px 0; }
                img { max-width: 100%; }
            </style>
        </head>
        <body>
            <h1>Report for {{ company_name }}</h1>
            
            {% for section in sections %}
            <div class="section">
                <h2>{{ section|replace('_', ' ')|title }}</h2>
                <!-- Add section content here -->
            </div>
            {% endfor %}
            
            {% if visualizations %}
            <h2>Visualizations</h2>
            {% for name, path in visualizations.items() %}
            <div class="visualization">
                <h3>{{ name|replace('_', ' ')|title }}</h3>
                <img src="{{ path }}" alt="{{ name }}">
            </div>
            {% endfor %}
            {% endif %}
        </body>
        </html>
        """
        return Template(template_str)

    def _add_table_of_contents(self, pdf, section_titles):
        """
        Add a professionally styled table of contents page with clickable links
        """
        # Add a new page for TOC
        pdf.add_page()
        
        # Modern TOC title with accent styling
        pdf.set_fill_color(*pdf.primary_color)
        pdf.rounded_rect(10, 40, 190, 15, 5, 'F')
        
        # Add accent bar on left
        pdf.set_fill_color(*pdf.accent_color)
        pdf.rounded_rect(10, 40, 5, 15, 5, 'F', corners='14')
        
        # TOC title
        pdf.set_font('Arial', 'B', 14)
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(20, 41.5)
        pdf.cell(170, 12, "TABLE OF CONTENTS", 0, 1, 'L')
        
        # Add decorative element
        pdf.set_draw_color(*pdf.accent_color)
        pdf.set_line_width(0.5)
        pdf.line(20, 65, 190, 65)
        
        # Add sections with modern styling
        y_position = 75
        for i, section_title in enumerate(section_titles):            
            # Calculate page number (cover + TOC + section pages)
            page_num = i + 3
            
            # Create a direct page link using FPDF's built-in linking system
            link_id = pdf.add_link()
            # This link will jump to the specified page
            pdf.set_link(link_id, page_num)
            
            # Section container with alternating background for better readability
            if i % 2 == 0:
                pdf.set_fill_color(250, 250, 252)  # Very light background
                pdf.rounded_rect(20, y_position - 3, 170, 12, 2, 'F')
            
            # Section number with accent circle
            pdf.set_fill_color(*pdf.accent_color)
            pdf.ellipse(25, y_position + 3, 3, 3, 'F')
            
            # Section title with link
            pdf.set_font('Arial', 'B', 11)
            pdf.set_text_color(*pdf.primary_color)
            pdf.set_xy(35, y_position)
            
            # Check if title needs to be truncated
            if pdf.get_string_width(section_title) > 110:
                title_truncated = section_title[:40] + "..."
                pdf.cell(110, 10, title_truncated, 0, 0, 'L', 0, link_id)
            else:
                pdf.cell(110, 10, section_title, 0, 0, 'L', 0, link_id)
            
            # Dotted line for visual connection to page number
            pdf.set_font('Arial', '', 10)
            pdf.set_text_color(150, 150, 150)
            
            dots_width = 30
            dots = '.' * int(dots_width / pdf.get_string_width('.'))
            pdf.cell(dots_width, 10, dots, 0, 0, 'C')
            
            # Page number with styled box (also clickable)
            pdf.set_fill_color(*pdf.primary_light)
            pdf.rounded_rect(175, y_position, 15, 8, 3, 'F')
            
            pdf.set_font('Arial', 'B', 9)
            pdf.set_text_color(*pdf.primary_color)
            pdf.set_xy(175, y_position)
            pdf.cell(15, 8, str(page_num), 0, 1, 'C', 0, link_id)
            
            # Move to next line
            y_position += 15
        
        # Add note at bottom
        note_y = max(210, y_position + 20)
        pdf.set_font('Arial', 'I', 9)
        pdf.set_text_color(100, 100, 100)
        pdf.set_xy(20, note_y)
        pdf.multi_cell(170, 5, "Click on any section title to navigate directly to that section. This report provides detailed financial analysis and risk assessment based on current market data and company performance metrics.", 0, 'C')
        
        # Return the starting page for the actual content
        return page_num + 1